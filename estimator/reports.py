from __future__ import annotations
from io import BytesIO
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from .models import ProjectInfo, MaterialLine, LaborLine, EstimateSettings, EstimateResult

CURRENCY = '$#,##0.00'

def materials_frame(lines: list[MaterialLine]) -> pd.DataFrame:
    return pd.DataFrame([{
        "Item": x.item, "Option": x.option, "Supplier": x.supplier,
        "Part Number": x.part_number, "Quantity": x.quantity,
        "Unit Cost": x.unit_cost, "Multiplier": x.multiplier,
        "Adjusted Unit Cost": x.adjusted_unit_cost, "Extended Cost": x.extended_cost,
    } for x in lines])

def labor_frame(lines: list[LaborLine]) -> pd.DataFrame:
    return pd.DataFrame([{
        "Category": x.category, "Hours": x.hours, "Hourly Rate": x.hourly_rate,
        "Burden Multiplier": x.burden_multiplier, "Extended Cost": x.extended_cost,
    } for x in lines])

def summary_frame(result: EstimateResult) -> pd.DataFrame:
    return pd.DataFrame([
        ("Material Base", result.material_base), ("Material Markup", result.material_markup),
        ("Taxable Material", result.taxable_material), ("Sales Tax", result.tax),
        ("Labor Base", result.labor_base), ("Labor Markup", result.labor_markup),
        ("Contingency", result.contingency), ("Overhead", result.overhead),
        ("Profit", result.profit), ("Grand Total", result.grand_total),
        ("Total Labor Hours", result.total_labor_hours), ("Estimated Duration (days)", result.duration_days),
    ], columns=["Metric", "Value"])

def build_excel(project, materials, labor, settings, result) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame({"Field": ["Project", "Customer", "Estimator", "Estimate Date", "County", "State", "Project Type", "Notes"],
                      "Value": [project.project_name, project.customer, project.estimator, project.estimate_date, project.county, project.state, project.project_type, project.notes]}).to_excel(writer, "Project", index=False)
        materials_frame(materials).to_excel(writer, "Materials", index=False)
        labor_frame(labor).to_excel(writer, "Labor", index=False)
        summary_frame(result).to_excel(writer, "Summary", index=False)
        wb = writer.book
        dark = PatternFill("solid", fgColor="1F4E78")
        accent = PatternFill("solid", fgColor="D9EAF7")
        white_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="B7B7B7")
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = dark; cell.font = white_font; cell.alignment = Alignment(horizontal="center")
            for column in ws.columns:
                width = min(max(len(str(c.value or "")) for c in column) + 2, 38)
                ws.column_dimensions[column[0].column_letter].width = width
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = Border(bottom=thin)
        for ws_name in ("Materials", "Labor", "Summary"):
            ws = wb[ws_name]
            for row in range(2, ws.max_row + 1):
                for cell in ws[row]:
                    if "Cost" in str(ws.cell(1, cell.column).value) or ws_name == "Summary" and cell.column == 2:
                        cell.number_format = CURRENCY
        ws = wb["Summary"]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value == "Grand Total":
                for cell in ws[row]: cell.fill = accent; cell.font = Font(bold=True)
    return output.getvalue()

def build_pdf(project, materials, labor, settings, result) -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=.45*inch, leftMargin=.45*inch, topMargin=.45*inch, bottomMargin=.45*inch)
    styles = getSampleStyleSheet()
    story = [Paragraph("Building Automation Estimate", styles["Title"]), Spacer(1, 10)]
    project_data = [["Project", project.project_name], ["Customer", project.customer], ["Estimator", project.estimator], ["Date", str(project.estimate_date)], ["Location", ", ".join(x for x in [project.county, project.state] if x)]]
    table = Table(project_data, colWidths=[1.2*inch, 5.8*inch]); table.setStyle(TableStyle([('GRID',(0,0),(-1,-1),.25,colors.grey),('BACKGROUND',(0,0),(0,-1),colors.HexColor('#D9EAF7'))])); story += [table, Spacer(1, 14)]
    story.append(Paragraph("Cost Summary", styles["Heading2"]))
    summary = [["Metric", "Value"]] + [[r.Metric, f"${r.Value:,.2f}" if r.Metric not in ("Total Labor Hours", "Estimated Duration (days)") else f"{r.Value:,.2f}"] for r in summary_frame(result).itertuples()]
    st = Table(summary, colWidths=[4.4*inch, 2.0*inch]); st.setStyle(TableStyle([('GRID',(0,0),(-1,-1),.25,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1F4E78')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('ALIGN',(1,1),(1,-1),'RIGHT'),('FONTNAME',(0,-3),(-1,-3),'Helvetica-Bold')])); story += [st, PageBreak()]
    story.append(Paragraph("Materials", styles["Heading2"]))
    mat_data = [["Item", "Part #", "Qty", "Unit", "Extended"]]
    for x in materials:
        mat_data.append([x.item[:42], x.part_number[:18], f"{x.quantity:g}", f"${x.adjusted_unit_cost:,.2f}", f"${x.extended_cost:,.2f}"])
    mt = Table(mat_data, repeatRows=1, colWidths=[2.7*inch,1.35*inch,.55*inch,.9*inch,1.05*inch]); mt.setStyle(TableStyle([('GRID',(0,0),(-1,-1),.2,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1F4E78')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTSIZE',(0,0),(-1,-1),7),('ALIGN',(2,1),(-1,-1),'RIGHT')])); story += [mt, Spacer(1,14), Paragraph("Labor", styles["Heading2"])]
    lab_data = [["Category", "Hours", "Rate", "Burden", "Extended"]] + [[x.category, f"{x.hours:g}", f"${x.hourly_rate:,.2f}", f"{x.burden_multiplier:.2f}", f"${x.extended_cost:,.2f}"] for x in labor]
    lt=Table(lab_data, repeatRows=1, colWidths=[2.8*inch,.75*inch,1*inch,.85*inch,1.15*inch]); lt.setStyle(TableStyle([('GRID',(0,0),(-1,-1),.2,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1F4E78')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('ALIGN',(1,1),(-1,-1),'RIGHT')])); story.append(lt)
    doc.build(story)
    return output.getvalue()
